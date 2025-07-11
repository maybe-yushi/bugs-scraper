from openpyxl import load_workbook
from bs4 import BeautifulSoup as bs
from urllib.request import urlopen


def main():
    workbook = load_workbook(filename="Music Releases.xlsx")
    # Select the active sheet
    sh_releases = workbook["Releases"]
    sh_artists = workbook["Artists"]
    empty_row = len([row for row in sh_releases if any(cell.value is not None for cell in row)]) + 1
    page = 1
    
    for row in sh_artists.iter_rows(min_row=2, max_row=len([row for row in sh_artists if any(cell.value is not None for cell in row)]), max_col=3):
        url_bugs_release = "https://music.bugs.co.kr/artist/" + str(row[0].value) + "/albums?type=RELEASE"
        webpage = urlopen(url_bugs_release)
        html = webpage.read().decode("utf-8")
        soup = bs(html, "html.parser")
        page = 1

        if soup.find("ul", {"class": "list tileView albumList"}) is not None:
            all_release = soup.find("ul", {"class": "list tileView albumList"}).find_all("li")
            number_release = len(all_release)
            while len(all_release) % 70 == 0:
                page += 1
                page_url = url_bugs_release + "&page=" + str(page)
                webpage = urlopen(page_url)
                html = webpage.read().decode("utf-8")
                soup = bs(html, "html.parser")
                all_release += soup.find("ul", {"class": "list tileView albumList"}).find_all("li")
                number_release = len(all_release)
            
            for rel in all_release:
                status = ""
                rel_id = rel.find("figure", {"class": "albumInfo"}).get("albumid").strip()
                for j in list(sh_releases.iter_rows(min_col=5, max_col=5, values_only=True)):
                    if j[0] == rel_id:
                        status = "already"
                        break

                if status != "already":
                    sh_releases.cell(row=empty_row, column=5).value = rel_id
                    rel_artist = row[1].value
                    sh_releases.cell(row=empty_row, column=1).value = rel_artist
                    rel_title = rel.find("div", {"class": "albumTitle"}).get_text().strip()
                    sh_releases.cell(row=empty_row, column=2).value = rel_title
                    rel_date = rel.find("time").get_text().strip()
                    sh_releases.cell(row=empty_row, column=3).value = rel_date
                    rel_type = rel.find("span", {"class": "albumType"}).get_text().strip()
                    sh_releases.cell(row=empty_row, column=4).value = rel_type
 
                    workbook.save("Music Releases.xlsx")
                    print("Album " + rel_id + " added.")
                    empty_row += 1
                    
            if int(row[2].value) != number_release:
                row[2].value = number_release
                print("Updated number of releases for " + str(row[1].value))
                workbook.save("Music Releases.xlsx")
        else:
            row[2].value = "Invalid ID"
            print("Artist " + str(row[1].value) + " has no releases.")
    print("Finished updating!")
    workbook.close()

if __name__ == '__main__':
    main()